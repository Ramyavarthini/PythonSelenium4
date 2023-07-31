#https://www.youtube.com/watch?v=S0WEaFucijs
import openpyxl
from openpyxl.styles import PatternFill

# get max row count
def get_rowcount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row

# get max col count
def get_col_count(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column

# read data
def read_data(file, sheetname,row,col):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row,col).value

# write data
def write_data(file,sheetname,row,col,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(row,col).value=data
    workbook.save(file)

#Fill Green colour
def fill_green_colour(file,sheetname,row,col):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    greenFill=PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(row,col).fill=greenFill
    workbook.save(file)

# Fill Red Colour
def fill_red_colour(file,sheetname,row,col):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    redFill=PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
    sheet.cell(row,col).fill=redFill
    workbook.save(file)

