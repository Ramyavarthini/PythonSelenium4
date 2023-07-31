# https://www.youtube.com/watch?v=S0WEaFucijs
# Install openpyxl in pycharm, File - settings - python Interpreter - '+' - 'openpyxl' - Install package
# Create an empty excel file
# close the specified excel file if it is opened.

import openpyxl

'''
file = "C:\SeleniumPracticeExcel\WritingSameData.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active  # or workbook["sheet1"] , botha are same if we have only one sheet in the file
for r in range(1,5):
    for c in range(1,4):
        sheet.cell(r,c).value="Ramya"
workbook.save(file)
'''
file = "C:\SeleniumPracticeExcel\WritingDiffData.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active
sheet.cell(1,1).value="101"
sheet.cell(1,2).value="Java"
sheet.cell(1,3).value="Ramya"

sheet.cell(2,1).value="201"
sheet.cell(2,2).value="Python"
sheet.cell(2,3).value="Ishana"

workbook.save(file)