# https://www.youtube.com/watch?v=S0WEaFucijs
# Install openpyxl in pycharm, File - settings - python Interpreter - '+' - 'openpyxl' - Install package
# close the specified excel file if it is opened.
import openpyxl

file = "C:\SeleniumPracticeExcel\BooksDeails.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
rows=sheet.max_row   # count number of rows in a excel sheet 5
cols=sheet.max_column  # count number of columns in a excel sheet 4
print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='        ')
    print()